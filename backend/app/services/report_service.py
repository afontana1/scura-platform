from __future__ import annotations

import csv
import html
import json
import tempfile
from datetime import datetime
from pathlib import Path
from uuid import UUID

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

from app.repositories.scenario_repository import ScenarioRepository
from app.repositories.simulation_repository import SimulationRepository
from app.schemas.scura_dataset import ScuraDataset, SimulationConfig


class ReportService:
    """Builds audit-ready report artifacts from completed simulation runs.

    The service uses canonical scenario JSON + stored simulation result JSON. It
    does not rerun the simulation and does not depend on frontend rendering.
    """

    def __init__(self, scenario_repo: ScenarioRepository, simulation_repo: SimulationRepository):
        self.scenario_repo = scenario_repo
        self.simulation_repo = simulation_repo

    def build_report_package(self, run_id: UUID, report_format: str) -> Path:
        bundle = self._load_completed_bundle(run_id)
        report_format = report_format.lower()
        out_dir = Path(tempfile.mkdtemp(prefix="scura_report_"))
        if report_format == "xlsx":
            return self._build_xlsx(bundle, out_dir / f"scura_report_{run_id}.xlsx")
        if report_format == "pdf":
            return self._build_pdf(bundle, out_dir / f"scura_report_{run_id}.pdf")
        if report_format == "html":
            return self._build_html(bundle, out_dir / f"scura_report_{run_id}.html")
        if report_format == "json":
            return self._build_json(bundle, out_dir / f"scura_report_{run_id}.json")
        if report_format == "csv":
            return self._build_csv_summary(bundle, out_dir / f"scura_report_{run_id}_summary.csv")
        raise HTTPException(status_code=404, detail="Unsupported report format")

    def get_report_manifest(self, run_id: UUID) -> dict:
        bundle = self._load_completed_bundle(run_id)
        summary = bundle["summary"]
        return {
            "run_id": str(run_id),
            "run_name": bundle["run"].run_name,
            "status": bundle["run"].status,
            "scenario_id": str(bundle["run"].scenario_id),
            "available_formats": ["pdf", "xlsx", "html", "json", "csv"],
            "summary_metrics": {
                "iterations": summary.get("iterations"),
                "engine_version": summary.get("engine_version"),
                "baseline_duration_days": summary.get("baseline_duration_days"),
                "baseline_cost": summary.get("baseline_cost"),
                "p80_duration_days": (summary.get("duration_percentiles") or {}).get("p80"),
                "p80_cost": (summary.get("cost_percentiles") or {}).get("p80"),
                "probability_meet_target_duration": summary.get("probability_meet_target_duration"),
                "probability_meet_target_budget": summary.get("probability_meet_target_budget"),
                "joint_probability": summary.get("joint_probability"),
            },
        }

    def _load_completed_bundle(self, run_id: UUID) -> dict:
        run = self.simulation_repo.get_run(run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Simulation run not found")
        if run.status != "completed":
            raise HTTPException(status_code=409, detail="Reports are only available for completed runs")
        result = self.simulation_repo.get_result_for_run(run_id)
        if not result:
            raise HTTPException(status_code=404, detail="Simulation result not found")
        dataset_record = self.scenario_repo.get_dataset(run.scenario_id)
        if not dataset_record:
            raise HTTPException(status_code=404, detail="Scenario dataset not found")
        dataset = ScuraDataset.model_validate(dataset_record.dataset_json)
        config = SimulationConfig.model_validate(run.config_json)
        return {
            "run": run,
            "dataset": dataset,
            "config": config,
            "summary": result.summary_json or {},
            "charts": result.charts_json or {},
            "drivers": result.driver_analysis_json or {},
            "artifacts": result.artifact_metadata_json or {},
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }

    def _flatten_summary(self, summary: dict) -> list[tuple[str, object]]:
        rows: list[tuple[str, object]] = []
        scalar_keys = [
            "engine_version", "iterations", "baseline_duration_days", "baseline_cost", "mean_duration_days",
            "mean_total_cost", "target_duration_days", "target_budget", "probability_meet_target_duration",
            "probability_meet_target_budget", "joint_probability", "schedule_contingency_p80", "cost_contingency_p80",
            "calendar_mode",
        ]
        for key in scalar_keys:
            if key in summary:
                rows.append((self._label(key), summary[key]))
        for parent in ["duration_percentiles", "cost_percentiles"]:
            values = summary.get(parent) or {}
            for key in sorted(values.keys(), key=lambda x: int(str(x).replace("p", "")) if str(x).replace("p", "").isdigit() else 999):
                rows.append((f"{self._label(parent)} {key.upper()}", values[key]))
        return rows

    def _build_xlsx(self, bundle: dict, path: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        header_fill = PatternFill("solid", fgColor="1F4E78")
        subheader_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(color="FFFFFF", bold=True)
        bold = Font(bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws["A1"] = "SCURA Simulation Report"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A2"] = "Run"
        ws["B2"] = bundle["run"].run_name
        ws["A3"] = "Generated At"
        ws["B3"] = bundle["generated_at"]
        ws["A4"] = "Scenario ID"
        ws["B4"] = str(bundle["run"].scenario_id)

        ws.append([])
        ws.append(["Metric", "Value"])
        for cell in ws[6]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        for metric, value in self._flatten_summary(bundle["summary"]):
            ws.append([metric, value])
        self._format_table(ws, 6, 1, ws.max_row, 2, border)
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 24

        self._add_rows_sheet(wb, "Risk Drivers", bundle["drivers"].get("risk_drivers") or [])
        self._add_rows_sheet(wb, "Criticality", bundle["drivers"].get("criticality_index") or [])
        self._add_rows_sheet(wb, "Activity Sensitivity", bundle["drivers"].get("activity_sensitivity") or [])
        self._add_rows_sheet(wb, "Milestone Confidence", bundle["drivers"].get("milestone_confidence") or [])
        self._add_chart_sheet(wb, "Duration S-Curve", bundle["charts"].get("duration_s_curve") or [])
        self._add_chart_sheet(wb, "Cost S-Curve", bundle["charts"].get("cost_s_curve") or [])

        assumptions = bundle["dataset"].assumptions or []
        self._add_rows_sheet(wb, "Assumptions", assumptions if isinstance(assumptions, list) else [])
        model_ws = wb.create_sheet("Model Configuration")
        model_ws.append(["Field", "Value"])
        for key, value in bundle["config"].model_dump(mode="json").items():
            model_ws.append([key, json.dumps(value) if isinstance(value, (list, dict)) else value])
        for cell in model_ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        model_ws.column_dimensions["A"].width = 34
        model_ws.column_dimensions["B"].width = 44

        for sheet in wb.worksheets:
            sheet.freeze_panes = "A2"
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        wb.save(path)
        return path

    def _add_rows_sheet(self, wb: Workbook, title: str, rows: list[dict]) -> None:
        ws = wb.create_sheet(title[:31])
        if not rows:
            ws.append(["No data"])
            return
        keys = sorted({key for row in rows for key in row.keys()})
        ws.append([self._label(key) for key in keys])
        for row in rows:
            ws.append([self._excel_value(row.get(key)) for key in keys])
        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        for idx, key in enumerate(keys, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = min(max(len(self._label(key)) + 4, 14), 42)

    def _add_chart_sheet(self, wb: Workbook, title: str, rows: list[dict]) -> None:
        ws = wb.create_sheet(title[:31])
        if not rows:
            ws.append(["No data"])
            return
        keys = sorted({key for row in rows for key in row.keys()})
        ws.append([self._label(key) for key in keys])
        for row in rows:
            ws.append([row.get(key) for key in keys])
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
        for idx in range(1, len(keys) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 18

    def _format_table(self, ws, min_row: int, min_col: int, max_row: int, max_col: int, border) -> None:
        for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = border

    def _build_pdf(self, bundle: dict, path: Path) -> Path:
        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(str(path), pagesize=landscape(LETTER), leftMargin=0.5 * inch, rightMargin=0.5 * inch, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        story = []
        story.append(Paragraph("SCURA Simulation Report", styles["Title"]))
        story.append(Paragraph(f"Run: {html.escape(bundle['run'].run_name)}", styles["Normal"]))
        story.append(Paragraph(f"Generated: {bundle['generated_at']}", styles["Normal"]))
        story.append(Spacer(1, 0.2 * inch))
        summary_rows = [["Metric", "Value"]] + [[metric, self._display(value)] for metric, value in self._flatten_summary(bundle["summary"])]
        story.append(self._pdf_table(summary_rows, [3.3 * inch, 2.6 * inch]))
        story.append(PageBreak())
        for title, rows in [
            ("Top Risk Drivers", bundle["drivers"].get("risk_drivers") or []),
            ("Criticality Index", bundle["drivers"].get("criticality_index") or []),
            ("Milestone Confidence", bundle["drivers"].get("milestone_confidence") or []),
            ("Activity Sensitivity", bundle["drivers"].get("activity_sensitivity") or []),
        ]:
            story.append(Paragraph(title, styles["Heading2"]))
            story.append(self._rows_to_pdf_table(rows[:15]))
            story.append(Spacer(1, 0.2 * inch))
        story.append(PageBreak())
        story.append(Paragraph("Model Configuration", styles["Heading2"]))
        config_rows = [[key, self._display(value)] for key, value in bundle["config"].model_dump(mode="json").items()]
        story.append(self._pdf_table([["Field", "Value"]] + config_rows, [2.8 * inch, 5.8 * inch]))
        doc.build(story)
        return path

    def _rows_to_pdf_table(self, rows: list[dict]) -> Table:
        if not rows:
            return self._pdf_table([["No data"]], [6 * inch])
        keys = sorted({key for row in rows for key in row.keys()})[:6]
        table_rows = [[self._label(key) for key in keys]]
        for row in rows:
            table_rows.append([self._display(row.get(key)) for key in keys])
        return self._pdf_table(table_rows, [1.35 * inch] * len(keys))

    def _pdf_table(self, rows: list[list[object]], col_widths: list[float]) -> Table:
        table = Table(rows, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9E2F3")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        return table

    def _build_html(self, bundle: dict, path: Path) -> Path:
        summary_rows = "".join(f"<tr><th>{html.escape(metric)}</th><td>{html.escape(self._display(value))}</td></tr>" for metric, value in self._flatten_summary(bundle["summary"]))
        risk_html = self._html_rows(bundle["drivers"].get("risk_drivers") or [])
        crit_html = self._html_rows(bundle["drivers"].get("criticality_index") or [])
        milestone_html = self._html_rows(bundle["drivers"].get("milestone_confidence") or [])
        content = f"""<!doctype html>
<html><head><meta charset='utf-8'><title>SCURA Report</title>
<style>body{{font-family:Arial,sans-serif;margin:32px;color:#172033}}table{{border-collapse:collapse;width:100%;margin:16px 0}}th,td{{border:1px solid #d9e2f3;padding:7px;text-align:left;vertical-align:top}}th{{background:#1f4e78;color:white}}h1,h2{{color:#1f4e78}}.meta{{color:#5b677a}}</style></head>
<body><h1>SCURA Simulation Report</h1><p class='meta'>Run: {html.escape(bundle['run'].run_name)}<br>Generated: {bundle['generated_at']}</p>
<h2>Executive Summary</h2><table>{summary_rows}</table>
<h2>Risk Drivers</h2>{risk_html}
<h2>Criticality Index</h2>{crit_html}
<h2>Milestone Confidence</h2>{milestone_html}
<h2>Model Configuration</h2><pre>{html.escape(json.dumps(bundle['config'].model_dump(mode='json'), indent=2))}</pre>
</body></html>"""
        path.write_text(content, encoding="utf-8")
        return path

    def _build_json(self, bundle: dict, path: Path) -> Path:
        payload = {
            "generated_at": bundle["generated_at"],
            "run": {"id": str(bundle["run"].id), "name": bundle["run"].run_name, "status": bundle["run"].status},
            "config": bundle["config"].model_dump(mode="json"),
            "summary": bundle["summary"],
            "charts": bundle["charts"],
            "driver_analysis": bundle["drivers"],
            "artifacts": bundle["artifacts"],
            "dataset_summary": {
                "activities": len(bundle["dataset"].schedule.activities),
                "relationships": len(bundle["dataset"].schedule.relationships),
                "cost_items": len(bundle["dataset"].cost.cost_items),
                "risk_events": len(bundle["dataset"].risks.risk_events),
                "risk_impacts": len(bundle["dataset"].risks.risk_impacts),
            },
        }
        path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        return path

    def _build_csv_summary(self, bundle: dict, path: Path) -> Path:
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["metric", "value"])
            for metric, value in self._flatten_summary(bundle["summary"]):
                writer.writerow([metric, self._display(value)])
        return path

    def _html_rows(self, rows: list[dict]) -> str:
        if not rows:
            return "<p>No data.</p>"
        keys = sorted({key for row in rows for key in row.keys()})[:8]
        head = "".join(f"<th>{html.escape(self._label(key))}</th>" for key in keys)
        body = "".join("<tr>" + "".join(f"<td>{html.escape(self._display(row.get(key)))}</td>" for key in keys) + "</tr>" for row in rows[:25])
        return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"

    def _label(self, key: str) -> str:
        return key.replace("_", " ").title()

    def _display(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            return f"{value:,.4f}" if abs(value) < 1 else f"{value:,.2f}"
        if isinstance(value, (list, dict)):
            return json.dumps(value)
        return str(value)

    def _excel_value(self, value: object) -> object:
        if isinstance(value, (list, dict)):
            return json.dumps(value)
        return value
