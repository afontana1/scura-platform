from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any
from uuid import UUID, uuid4

from openpyxl import load_workbook

from app.schemas.imports import ImportPreview
from app.schemas.scenario import ScenarioDatasetUpsert
from app.services.scenario_service import ScenarioService
from app.services.validation_service import ValidationService


@dataclass
class StagedImport:
    import_id: UUID
    filename: str
    dataset_json: dict[str, Any]
    sheet_row_counts: dict[str, int]


_STAGED_IMPORTS: dict[UUID, StagedImport] = {}


class ImportService:
    """Converts spreadsheet uploads into canonical SCURA JSON.

    Excel is treated as an import adapter only. The rest of the backend works
    with canonical SCURA JSON and Pydantic models.
    """

    def __init__(self, validation_service: ValidationService):
        self.validation_service = validation_service

    def parse_excel(self, filename: str, content: bytes) -> ImportPreview:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
        tables = {sheet_name: self._sheet_to_rows(workbook[sheet_name]) for sheet_name in workbook.sheetnames}
        dataset = self._tables_to_dataset(tables)
        sheet_row_counts = {name: len(rows) for name, rows in tables.items() if rows}
        staged = StagedImport(uuid4(), filename, dataset, sheet_row_counts)
        _STAGED_IMPORTS[staged.import_id] = staged
        validation = self.validation_service.validate_dataset_payload(dataset)
        return ImportPreview(
            import_id=staged.import_id,
            filename=filename,
            dataset_json=dataset,
            validation=validation,
            sheet_row_counts=sheet_row_counts,
        )

    def get_preview(self, import_id: UUID) -> ImportPreview:
        staged = self._get_staged(import_id)
        validation = self.validation_service.validate_dataset_payload(staged.dataset_json)
        return ImportPreview(
            import_id=staged.import_id,
            filename=staged.filename,
            dataset_json=staged.dataset_json,
            validation=validation,
            sheet_row_counts=staged.sheet_row_counts,
        )

    def commit(self, import_id: UUID, scenario_service: ScenarioService, scenario_id: UUID):
        staged = self._get_staged(import_id)
        validation = self.validation_service.validate_dataset_payload(staged.dataset_json)
        payload = ScenarioDatasetUpsert(
            dataset_json=staged.dataset_json,
            schema_version=str(staged.dataset_json.get("schema_version", "0.1.0")),
            is_valid=validation.valid,
        )
        return scenario_service.upsert_dataset(scenario_id, payload)

    def _get_staged(self, import_id: UUID) -> StagedImport:
        staged = _STAGED_IMPORTS.get(import_id)
        if staged is None:
            raise KeyError(f"Import {import_id} was not found. Upload the workbook again.")
        return staged

    def _sheet_to_rows(self, worksheet) -> list[dict[str, Any]]:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [self._normalize_header(value) for value in rows[0]]
        output: list[dict[str, Any]] = []
        for row in rows[1:]:
            item: dict[str, Any] = {}
            has_value = False
            for header, value in zip(headers, row):
                if not header:
                    continue
                normalized = self._clean_cell(value)
                if normalized not in (None, ""):
                    has_value = True
                item[header] = normalized
            if has_value:
                output.append(item)
        return output

    def _tables_to_dataset(self, tables: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
        activities = [self._activity(row) for row in self._table(tables, "Schedule_Activities")]
        relationships = [self._relationship(row, index) for index, row in enumerate(self._table(tables, "Schedule_Relationships"), start=1)]
        if not relationships:
            relationships = self._relationships_from_predecessors(self._table(tables, "Schedule_Activities"))

        dataset = {
            "schema_version": "0.3.0",
            "project": {"project_id": None, "name": None},
            "scenario": {"scenario_id": None, "name": None, "status_date": None},
            "schedule": {
                "activities": activities,
                "relationships": relationships,
                "duration_uncertainties": [self._duration_uncertainty(row) for row in self._table(tables, "Schedule_Uncertainty")],
                "calendars": [self._calendar(row) for row in self._table(tables, "Calendars")] or [{"calendar_id": "standard_5_day", "name": "Standard 5-Day", "workdays_per_week": 5, "notes": None}],
                "milestones": [self._milestone(row, index) for index, row in enumerate(self._table(tables, "Milestones"), start=1)],
            },
            "cost": {
                "cost_items": [self._cost_item(row) for row in self._table(tables, "Cost_Estimate")],
                "cost_uncertainties": [self._cost_uncertainty(row) for row in self._table(tables, "Cost_Uncertainty")],
                "cost_schedule_mappings": [self._cost_mapping(row, index) for index, row in enumerate(self._table(tables, "Cost_Schedule_Map"), start=1)],
            },
            "risks": {
                "risk_events": [self._risk_event(row) for row in self._table(tables, "Risk_Register")],
                "risk_impacts": [self._risk_impact(row, index) for index, row in enumerate(self._table(tables, "Risk_Impacts"), start=1)],
            },
            "correlations": [self._correlation(row, index) for index, row in enumerate(self._table(tables, "Correlations"), start=1)],
            "assumptions": self._table(tables, "Assumptions"),
        }
        return dataset

    def _table(self, tables: dict[str, list[dict[str, Any]]], sheet_name: str) -> list[dict[str, Any]]:
        normalized_target = self._normalize_header(sheet_name)
        for name, rows in tables.items():
            if self._normalize_header(name) == normalized_target:
                return rows
        return []

    def _activity(self, row: dict[str, Any]) -> dict[str, Any]:
        baseline = self._number(row.get("baseline_duration_days") or row.get("baseline_duration") or row.get("duration"), 0)
        remaining = row.get("remaining_duration_days")
        return {
            "activity_id": self._text(row.get("activity_id")),
            "name": self._text(row.get("name") or row.get("activity_name")),
            "wbs_id": self._nullable_text(row.get("wbs_id") or row.get("wbs")),
            "baseline_duration_days": baseline,
            "calendar_id": self._nullable_text(row.get("calendar_id") or row.get("calendar")) or "standard_5_day",
            "status": self._text(row.get("status") or "not_started"),
            "actual_start": self._date_text(row.get("actual_start")),
            "actual_finish": self._date_text(row.get("actual_finish")),
            "remaining_duration_days": self._number(remaining, baseline) if remaining not in (None, "") else baseline,
        }

    def _relationship(self, row: dict[str, Any], index: int) -> dict[str, Any]:
        return {
            "relationship_id": self._text(row.get("relationship_id") or f"REL-{index:04d}"),
            "predecessor_activity_id": self._text(row.get("predecessor_activity_id") or row.get("predecessor")),
            "successor_activity_id": self._text(row.get("successor_activity_id") or row.get("successor")),
            "relationship_type": self._text(row.get("relationship_type") or row.get("type") or "FS"),
            "lag_days": self._number(row.get("lag_days") or row.get("lag"), 0),
        }

    def _relationships_from_predecessors(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        relationships = []
        counter = 1
        for row in rows:
            successor = self._text(row.get("activity_id"))
            predecessors = self._text(row.get("predecessors") or "")
            for predecessor in [item.strip() for item in predecessors.split(",") if item.strip()]:
                relationships.append({
                    "relationship_id": f"REL-{counter:04d}",
                    "predecessor_activity_id": predecessor,
                    "successor_activity_id": successor,
                    "relationship_type": "FS",
                    "lag_days": 0,
                })
                counter += 1
        return relationships


    def _calendar(self, row: dict[str, Any]) -> dict[str, Any]:
        return {
            "calendar_id": self._text(row.get("calendar_id") or row.get("id") or "standard_5_day"),
            "name": self._text(row.get("name") or row.get("calendar_name") or row.get("calendar_id") or "Standard 5-Day"),
            "workdays_per_week": self._number(row.get("workdays_per_week") or row.get("workdays") or 5, 5),
            "notes": self._nullable_text(row.get("notes")),
        }

    def _milestone(self, row: dict[str, Any], index: int) -> dict[str, Any]:
        target = row.get("target_day") or row.get("target_duration_days")
        return {
            "milestone_id": self._text(row.get("milestone_id") or f"MS-{index:04d}"),
            "name": self._text(row.get("name") or row.get("milestone_name") or f"Milestone {index}"),
            "activity_id": self._text(row.get("activity_id")),
            "target_day": self._number(target, 0) if target not in (None, "") else None,
        }

    def _correlation(self, row: dict[str, Any], index: int) -> dict[str, Any]:
        raw_targets = self._text(row.get("target_ids") or row.get("targets") or "")
        target_ids = [item.strip() for item in raw_targets.split(",") if item.strip()]
        return {
            "correlation_id": self._text(row.get("correlation_id") or f"CORR-{index:04d}"),
            "name": self._text(row.get("name") or row.get("correlation_name") or f"Correlation {index}"),
            "target_type": self._text(row.get("target_type") or "activity_duration"),
            "target_ids": target_ids,
            "strength": self._number(row.get("strength"), 0.5),
            "notes": self._nullable_text(row.get("notes")),
        }

    def _duration_uncertainty(self, row: dict[str, Any]) -> dict[str, Any]:
        return self._uncertainty(row, "activity_id")

    def _cost_uncertainty(self, row: dict[str, Any]) -> dict[str, Any]:
        return self._uncertainty(row, "cost_id")

    def _uncertainty(self, row: dict[str, Any], id_field: str) -> dict[str, Any]:
        return {
            id_field: self._text(row.get(id_field)),
            "distribution": self._text(row.get("distribution") or "triangular"),
            "minimum": self._number(row.get("minimum") or row.get("min"), 0),
            "most_likely": self._number(row.get("most_likely") or row.get("likely") or row.get("mode"), 0),
            "maximum": self._number(row.get("maximum") or row.get("max"), 0),
        }

    def _cost_item(self, row: dict[str, Any]) -> dict[str, Any]:
        return {
            "cost_id": self._text(row.get("cost_id")),
            "wbs_id": self._nullable_text(row.get("wbs_id") or row.get("wbs")),
            "description": self._text(row.get("description") or row.get("cost_description")),
            "baseline_cost": self._number(row.get("baseline_cost") or row.get("cost"), 0),
            "cost_type": self._text(row.get("cost_type") or "fixed"),
            "currency": self._text(row.get("currency") or "USD"),
        }

    def _cost_mapping(self, row: dict[str, Any], index: int) -> dict[str, Any]:
        return {
            "mapping_id": self._text(row.get("mapping_id") or f"MAP-{index:04d}"),
            "cost_id": self._text(row.get("cost_id")),
            "activity_id": self._text(row.get("activity_id")),
            "behavior": self._text(row.get("behavior") or "scale_with_duration"),
        }

    def _risk_event(self, row: dict[str, Any]) -> dict[str, Any]:
        return {
            "risk_id": self._text(row.get("risk_id")),
            "name": self._text(row.get("name") or row.get("risk_name")),
            "description": self._nullable_text(row.get("description")),
            "probability": self._number(row.get("probability"), 0),
            "owner": self._nullable_text(row.get("owner")),
            "status": self._text(row.get("status") or "active"),
        }

    def _risk_impact(self, row: dict[str, Any], index: int) -> dict[str, Any]:
        min_delay = row.get("min_delay") or row.get("minimum_delay")
        likely_delay = row.get("most_likely_delay") or row.get("likely_delay")
        max_delay = row.get("max_delay") or row.get("maximum_delay")
        min_cost = row.get("min_cost") or row.get("minimum_cost")
        likely_cost = row.get("most_likely_cost") or row.get("likely_cost")
        max_cost = row.get("max_cost") or row.get("maximum_cost")
        item: dict[str, Any] = {
            "impact_id": self._text(row.get("impact_id") or f"RI-{index:04d}"),
            "risk_id": self._text(row.get("risk_id")),
            "activity_id": self._nullable_text(row.get("activity_id")),
            "cost_id": self._nullable_text(row.get("cost_id")),
            "schedule_impact": None,
            "cost_impact": None,
        }
        if any(value not in (None, "") for value in [min_delay, likely_delay, max_delay]):
            item["schedule_impact"] = {
                "distribution": self._text(row.get("schedule_distribution") or row.get("distribution") or "triangular"),
                "minimum": self._number(min_delay, 0),
                "most_likely": self._number(likely_delay, 0),
                "maximum": self._number(max_delay, 0),
            }
        if any(value not in (None, "") for value in [min_cost, likely_cost, max_cost]):
            item["cost_impact"] = {
                "distribution": self._text(row.get("cost_distribution") or row.get("distribution") or "triangular"),
                "minimum": self._number(min_cost, 0),
                "most_likely": self._number(likely_cost, 0),
                "maximum": self._number(max_cost, 0),
            }
        return item

    def _normalize_header(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip().lower().replace(" ", "_").replace("-", "_")

    def _clean_cell(self, value: Any) -> Any:
        if isinstance(value, str):
            return value.strip()
        return value

    def _text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _nullable_text(self, value: Any) -> str | None:
        text = self._text(value)
        return text if text else None

    def _number(self, value: Any, default: float) -> float:
        if value in (None, ""):
            return float(default)
        try:
            return float(value)
        except (TypeError, ValueError):
            return float(default)

    def _date_text(self, value: Any) -> str | None:
        if value in (None, ""):
            return None
        if hasattr(value, "date"):
            return value.date().isoformat()
        return self._text(value)
